# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter


class SpiderdemoPipeline:
	def process_item(self, item, spider):
		return item

import openpyxl
from .items import BookItem
class BookItemPipeline:
	def open_spider(self, spider):
		# 可选实现，开启spider时调用该方法
		self.wb = openpyxl.Workbook()
		self.sheet = self.wb.active
		self.sheet.title = 'BooksMessage'
		self.sheet.append(('作者', '书名', '价格'))

	def process_item(self, item: BookItem, spider):
		self.sheet.append((item['author'], item['name'],  item['price']))
		return item

	def close_spider(self, spider):
		self.wb.save('异步社区图书信息.xlsx')



'''
from .items import DoubanItem


class MovieItemPipeline:

	# def __init__(self):
	#     self.wb = openpyxl.Workbook()
	#     self.sheet = self.wb.active
	#     self.sheet.title = 'Top250'
	#     self.sheet.append(('名称', '评分', '名言'))

	def open_spider(self, spider):
		#可选实现，开启spider时调用该方法
		self.wb = openpyxl.Workbook()
		self.sheet = self.wb.active
		self.sheet.title = 'Top250'
		self.sheet.append(('名称', '评分', '名言'))

	def process_item(self, item: DoubanItem, spider):
		self.sheet.append((item['title'], item['score'], item['motto']))
		return item

	def close_spider(self, spider):
		self.wb.save('豆瓣电影数据.xlsx')
'''