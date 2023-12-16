# -*- coding:utf-8 -*-

# @Time: 2023/12/16 14:11
# @Project: test
# @File: dataPre.py
# @Author: WSM
import numpy as np
import pandas as pd
import jsonpath
import json
import random

from openpyxl.reader.excel import load_workbook

# 示例 JSON 字符串
json_str = '''  
{
    "param":{
        "accessChannel":"SDK",
        "businessMode":6,
        "businessScene":601,
        "buyer":1239095,
        "executionEnvType":"ANDROID",
        "flowNo":"62023101711002${sys.random(100000000000,999990000000)}",
        "orderLines":[
            {
                "commodityId":200004983,
                "liveCategoryId":106,
                "quantity":1,
                "receiver":1243048
            }
        ],
        "orderSource":{
            "adminerUid":1243048,
            "appSource":"XIMA",
            "liveRecordId":1078086,
            "liveSource":"PERSONAL_LIVE",
            "locationSource":"LIVE_ROOM",
            "mediaType":"VIDEO",
            "roomId":840056,
            "roomOwnerUid":1243048
        },
        "payInfo":{
            "accountType":"XI_BEAN",
            "combinedPayType":"COMBINED_PAY_PRIOR_DEDUCT_NOBLE_BEAN",
            "payMode":"BALANCE",
            "subAccountType":100,
            "userId":1239095
        },
        "payee":-11000,
        "title":"(压测)用户购买立即赠送：送1个礼物给主播",
        "tradeMode":"USER_BUY_AND_SEND",
        "tradeTime":1692172753060,
        "tradeType":"BUY_AUTO_DELIVER"
    }
}
'''
def read_xlsx(f):
    # 存放表位置
    file= f
    book = load_workbook(file)
    sheet = book['Sheet1']
    # 获取总行数
    rows = sheet.max_row
    # 获取总列数
    # cols = sheet.max_col
    # 获取表头
    head = [row for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0]
    # 数据组装
    datas = []
    for row in sheet.iter_rows(min_row=2, max_row=rows + 1, values_only=True):
        data = dict(zip(head, row))
        datas.append(data)
    return datas

# 定义一个函数用于生成新的 JSON 字符串
def generate_new_json_string(json_str, new_user_id,new_roomOwnerUid):

    json_str1 = json_str.replace('"buyer":1239095', f'"buyer": {new_user_id}')
    json_str2 = json_str1.replace('"adminerUid":1243048', f'"adminerUid": {new_roomOwnerUid}')
    json_str3 = json_str2.replace('"receiver":1243048', f'"receiver": {new_roomOwnerUid}')
    json_str4 = json_str3.replace('"roomOwnerUid":1243048', f'"roomOwnerUid": {new_roomOwnerUid}')

    # json_str5 = json_str4.replace('"commodityId":200005199', f'"commodityId": {commodityId}')

    new_json_str = json_str4.replace('"userId":1239095', f'"userId": {new_user_id}')

    return new_json_str


# 创建一个空的 DataFrame 用于存储新生成的 JSON 字符串
new_df = pd.DataFrame(columns=['json_str'])
json_list = []

datas = read_xlsx('uid.xlsx')
# datas = read_xlsx('uid-a.xlsx')

# datas = read_xlsx('uid-b.xlsx')

# datas2 = read_xlsx('xx.xlsx')

for i in range(0,3000):
    # print("对应userid")
    new_user_id = datas[i]['user_id']
    # print(new_user_id)

    # print("对应roomOwnerUid")
    new_roomOwnerUid = datas[i+1]['user_id']
    # print(new_roomOwnerUid)

    # commodityId = datas2[i]['']

    # print("新的json")
    new_json_str = generate_new_json_string(json_str, new_user_id, new_roomOwnerUid)
    # new_json_str = generate_new_json_string(json_str, new_user_id, new_roomOwnerUid,commodityId)
    # print(new_json_str)

    new_df = new_df.append({'json_str': new_json_str}, ignore_index=True)

    json_list.append(json.loads(new_json_str))

# 将新生成的 JSON 字符串保存到新的 Excel 文件中
new_df.to_excel('consume_1.xlsx', index=False)
# print("json列表")
print(json_list)

# 将列表转换为 JSON 字符串
json1 = json.dumps(json_list)
# 将 JSON 字符串写入文件
with open('consume_1.json', 'w') as f:
    f.write(json1)
