# -*- coding: utf-8 -*-

# @Project : apiTest
# @File    : myexcel.py
# @Date    : 2019-11-29
# @Author  : hutong
# @Describe: 微信公众： 大话性能


import openpyxl

from logUtil.mylog import Log
mylogger = Log(__name__).getlogger()

class OperateExcel():
    def __init__(self, excelpath, sheetname):
        self.filename = excelpath
        self.wb = openpyxl.load_workbook(excelpath) #加载已经存在的excel文件，后缀名要xlsx
        self.sh = self.wb[sheetname]

    #按行读取指定sheet表单中所有的内容
    def read_all_data_line_by_line(self):
        """一行一行的获取数据"""
        row_datas = list(self.sh.rows)  # 按行获取数据转换成列表
        titles = []  # 获取表单的表头信息
        for title in row_datas[0]:  # 获取第一行的数据，row_datas从下标为0开始
            titles.append(title.value)
        testdatas = []  # 存储所有行数据
        for case in row_datas[1:]:  # 从第二行开始开始获取数据，case数据类型：元组
            data = []  # 临时存储每一行的数据
            for cell in case:
                try:
                    #data.append(eval(cell.value))
                    data.append(cell.value)
                except Exception as e:
                    #data.append(cell.value)
                    mylogger.error('读取excel行数据失败：{}'.format(e),exc_info=True)
            case_data = dict(list(zip(titles, data)))  # case_data存储的是一行的数据，存储格式：{title:cell_value}
            testdatas.append(case_data)
        return testdatas

    # 按行读取指定sheet表单中指定列中的内容
    def read_all_data_column_by_column(self, columns):
            """获取指定几列的信息
            columns可以是列表，也可以是元组类型，如[1,2,3]就代表获取第1，2，3列的数据"""
            maxline = self.sh.max_row  # 获取最大行数
            titles = []  # 存储标题
            alldatas = []
            for linenum in range(1, maxline + 1):  # 从第一行开始，range属于左闭右开，所以右侧+1
                if linenum != 1:  # 如果是第一行的话，追加到titles中
                    onelinedata = []  # 一行数据
                    for column in columns:  # 遍历想要获取的列数
                        one_cell_value = self.sh.cell(linenum, column).value
                        try:
                            onelinedata.append(eval(one_cell_value))  # 如果可转为dict,tuple，list，直接转型，再取出来
                        except Exception as e:
                            onelinedata.append(one_cell_value)
                    onelinedata_dict = dict(list(zip(titles, onelinedata)))
                    alldatas.append(onelinedata_dict)

                else:
                    for column in columns:
                        titles.append(self.sh.cell(linenum, column).value)

            return alldatas
    # 向指定单元格中写入数据
    def write_content_to_row_column(self, row, column, content):
        """
        1.  row 和 column代表单元格所在行和列
        2.  content 是写入内容，类型是字符串"""
        try:
            self.sh.cell(row=row, column=column, value=content)
            self.wb.save('test2.xlsx')  # 保存文件
        except Exception as e:
            mylogger.error('写入excel数据失败：{}'.format(e), exc_info=True)



if __name__ == "__main__":
    filepath = '/PycharmProjects/apiTest/test_case_data/socketcase.xlsx'  # 文件绝对路径,也可以是相对路径
    sheetname = 'hw'  # sheetname
    sheetObject = OperateExcel(filepath, sheetname)
    testdatas = sheetObject.read_all_data_line_by_line()
    for onetestdata in testdatas:
        #print(onetestdata)
        mylogger.info(onetestdata)
        #sheetObject.write_content_to_row_column(7,1,'test')

